from io import BytesIO

import openpyxl
import pytest
from django.conf import settings

from fnac.models.new_product_export import create_new_product_export


@pytest.fixture
def export_file():
    export_file = create_new_product_export()
    return openpyxl.load_workbook(filename=BytesIO(export_file.read()))


@pytest.fixture
def product(fnac_product_factory, translation_factory):
    product = fnac_product_factory.create()
    translation_factory.create(product=product)
    return product


@pytest.fixture
def export_row(product, export_file):
    return [_.value for _ in list(export_file.active.rows)[2]]


@pytest.fixture
def invalid_product(fnac_product_factory):
    return fnac_product_factory.create()


@pytest.mark.django_db
def test_response_header(product, export_file):
    worksheet = export_file.active
    assert list(worksheet.rows)[0][0].value == "Typologie"
    assert list(worksheet.rows)[1][0].value == "Typology"


@pytest.mark.django_db
@pytest.mark.parametrize(
    "product_kwargs", [{}, {"english_size": "", "french_size": None}]
)
def test_response_row(product_kwargs, fnac_product_factory, translation_factory):
    product = fnac_product_factory.create(**product_kwargs)
    translation_factory.create(product=product)
    export_file = create_new_product_export()
    workbook = openpyxl.load_workbook(filename=BytesIO(export_file.read()))
    export_row = [_.value for _ in list(workbook.active.rows)[2]]
    if product.french_size is None:
        french_size = None
    else:
        french_size = product.french_size.name
    cell_values = [
        product.sku,
        product.translation.name,
        product.translation.description,
        product.translation.description,
        "Accessoires",
        product.fnac_range.category.french,
        product.barcode,
        product.brand,
        product.translation.colour,
        french_size,
        settings.CC_IMAGE_URL + product.image_1,
        settings.CC_IMAGE_URL + product.image_2,
        settings.CC_IMAGE_URL + product.image_3,
    ]
    for value in cell_values:
        assert value in export_row


@pytest.mark.django_db
@pytest.mark.parametrize(
    "name,expected",
    [
        ("Non Matching", "Accessoires"),
        ("plush", "Peluche"),
        ("sticker", "Autocollant / Magnets"),
    ],
)
def test_format_is_set(name, expected, fnac_product_factory, translation_factory):
    product = fnac_product_factory(name=name)
    translation_factory.create(product=product)
    export_file = create_new_product_export()
    worksheet = openpyxl.load_workbook(filename=BytesIO(export_file.read())).active
    row = [_.value for _ in list(worksheet.rows)[2]]
    assert expected in row


@pytest.mark.django_db
def test_invaild_product_is_not_exported(
    product, invalid_product, export_file, export_row
):
    assert len(list(export_file.active.rows)) == 3
    assert invalid_product.sku not in export_row
