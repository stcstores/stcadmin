import openpyxl
import pytest

from fnac.models.add_missing_information import (
    _ImportMissingInformationFile,
    import_missing_information,
)

HEADER = _ImportMissingInformationFile.HEADER


@pytest.fixture
def empty_import_row():
    return [None] * len(HEADER)


@pytest.fixture
def size(size_factory):
    return size_factory.create(name="XL")


@pytest.fixture
def category(category_factory):
    return category_factory.create(
        name="category_name", english="category_english", french="category_french"
    )


@pytest.fixture
def import_row(empty_import_row):
    def _import_row(
        sku="ABC-DEF-123", colour=None, category=None, size=None, price=None
    ):
        row = empty_import_row
        row[HEADER.index(_ImportMissingInformationFile.SKU)] = sku
        row[HEADER.index(_ImportMissingInformationFile.COLOUR)] = colour
        row[HEADER.index(_ImportMissingInformationFile.CATEGORY)] = category
        row[HEADER.index(_ImportMissingInformationFile.SIZE_FRENCH)] = size
        row[HEADER.index(_ImportMissingInformationFile.PRICE)] = price
        return row

    return _import_row


@pytest.fixture
def expected_update_kwargs(size, category):
    def _expected_update_kwargs(kwargs):
        if kwargs.get("french_size") == "SIZE":
            kwargs["french_size"] = size
        if kwargs.get("category") == "CATEGORY":
            kwargs["category"] = category
        return kwargs

    return _expected_update_kwargs


@pytest.fixture
def import_file(tmpdir, import_row):
    def _import_file(row):
        row = import_row(**row)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = _ImportMissingInformationFile.PRODUCTS_SHEET
        for column_number, header in enumerate(HEADER):
            worksheet.cell(column=column_number + 1, row=1).value = header
            worksheet.cell(column=column_number + 1, row=2).value = row[column_number]
        path = tmpdir.join("import_file.xlsx")
        with open(path, "wb") as f:
            workbook.save(f)
        return path

    return _import_file


@pytest.mark.django_db
@pytest.mark.parametrize(
    "import_data,expected",
    (
        ({}, {}),
        ({"size": "XL"}, {"french_size": "SIZE"}),
        ({"size": "NA"}, {}),
        ({"colour": "Red"}, {"colour": "Red"}),
        ({"colour": ""}, {}),
        ({"price": 5.99}, {"price": 599}),
        ({"category": "category_name"}, {"category": "CATEGORY"}),
        ({"category": "category_english"}, {"category": "CATEGORY"}),
        ({"category": "category_french"}, {"category": "CATEGORY"}),
        ({"category": "INVALID"}, {}),
        (
            {
                "category": "category_name",
                "price": 6.89,
                "size": "XL",
                "colour": "Blue",
            },
            {
                "category": "CATEGORY",
                "french_size": "SIZE",
                "price": 689,
                "colour": "Blue",
            },
        ),
    ),
)
def test_update_kwargs(import_data, expected, import_row, expected_update_kwargs):
    row = import_row(**import_data)
    expected = expected_update_kwargs(expected)
    assert _ImportMissingInformationFile().get_update_kwargs(row) == expected


@pytest.mark.django_db
def test_category_updated(fnac_product_factory, import_file, category):
    product = fnac_product_factory.create()
    workbook = import_file({"sku": product.sku, "category": category.name})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.fnac_range.category == category


@pytest.mark.django_db
def test_category_does_not_change_if_invalid(
    fnac_product_factory, import_file, category
):
    product = fnac_product_factory.create()
    original_category = product.fnac_range.category
    workbook = import_file({"sku": product.sku, "category": "invalid"})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.fnac_range.category == original_category


@pytest.mark.django_db
def test_colour_updated(fnac_product_factory, import_file):
    product = fnac_product_factory.create()
    workbook = import_file({"sku": product.sku, "colour": "Magenta"})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.colour == "Magenta"


@pytest.mark.django_db
def test_colour_not_updated_if_empty(fnac_product_factory, import_file):
    product = fnac_product_factory.create(colour="Orange")
    workbook = import_file({"sku": product.sku, "colour": ""})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.colour == "Orange"


@pytest.mark.django_db
def test_price_updated(fnac_product_factory, import_file):
    product = fnac_product_factory.create()
    workbook = import_file({"sku": product.sku, "price": 12.59})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.price == 1259


@pytest.mark.django_db
def test_price_not_updated_if_empty(fnac_product_factory, import_file):
    product = fnac_product_factory.create(price=1548)
    workbook = import_file({"sku": product.sku, "price": None})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.price == 1548


@pytest.mark.django_db
def test_size_updated(fnac_product_factory, import_file, size):
    product = fnac_product_factory.create(french_size=None)
    workbook = import_file({"sku": product.sku, "size": size.name})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.french_size == size


@pytest.mark.django_db
def test_size_not_updated_if_empty(fnac_product_factory, import_file, size):
    product = fnac_product_factory.create(french_size=size)
    workbook = import_file({"sku": product.sku, "size": None})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.french_size == size


@pytest.mark.django_db
def test_size_not_updated_if_invalid(fnac_product_factory, import_file, size):
    product = fnac_product_factory.create(french_size=size)
    workbook = import_file({"sku": product.sku, "size": "INVALID SIZE"})
    import_missing_information(workbook)
    product.refresh_from_db()
    assert product.french_size == size


@pytest.mark.django_db
def test_unrecognised_products_are_ignored(import_file, category, size):
    workbook = import_file(
        {
            "sku": "INVALID_SKU",
            "category": category.name,
            "size": "XL",
            "price": 10.58,
            "colour": "Red",
        }
    )
    import_missing_information(workbook)
