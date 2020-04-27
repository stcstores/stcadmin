import openpyxl
import pytest

from fnac.models.add_missing_information import (
    _MissingInformationExportFile,
    create_add_missing_information_export,
)


@pytest.fixture
def export_file(db):
    binary_file = create_add_missing_information_export()
    return openpyxl.load_workbook(binary_file, data_only=True)


@pytest.fixture
def sizes(size_factory):
    return [size_factory.create() for _ in range(5)]


@pytest.fixture
def colour_required_categories(category_factory):
    return [category_factory.create(requires_colour=True) for i in range(5)]


@pytest.fixture
def categories(category_factory, colour_required_categories):
    return [category_factory.create() for i in range(5)] + colour_required_categories


@pytest.mark.django_db
@pytest.mark.parametrize(
    "sheet_name",
    [
        _MissingInformationExportFile.PRODUCTS_SHEET,
        _MissingInformationExportFile.CATEGORIES_SHEET,
        _MissingInformationExportFile.SIZES_SHEET,
        _MissingInformationExportFile.INSTRUCTIONS_SHEET,
        _MissingInformationExportFile.COLOUR_REQUIRED_SHEET,
    ],
)
def test_export_file_has_sheets(export_file, sheet_name):
    assert sheet_name in export_file.sheetnames


def test_instructions_sheet(export_file):
    instructions = _MissingInformationExportFile.INSTRUCTIONS
    instruction_rows = [
        [col, instructions[col]] for col in _MissingInformationExportFile.HEADER
    ]
    worksheet = export_file[_MissingInformationExportFile.INSTRUCTIONS_SHEET]
    worksheet_rows = list(worksheet.rows)
    header = worksheet_rows[0]
    assert [header[0].value, header[1].value] == [
        "Column",
        "Use",
    ]
    for i, row in enumerate(worksheet_rows[1:]):
        assert [row[0].value, row[1].value] == instruction_rows[i]


@pytest.mark.django_db
def test_sizes_sheet(sizes, export_file):
    worksheet = export_file[_MissingInformationExportFile.SIZES_SHEET]
    worksheet_rows = list(worksheet.rows)
    for i, size in enumerate(sizes):
        assert worksheet_rows[i][0].value == size.name


@pytest.mark.django_db
def test_colour_categories_sheet(categories, colour_required_categories, export_file):
    category_rows = [[_.name, _.english, _.french] for _ in colour_required_categories]
    worksheet_rows = list(
        export_file[_MissingInformationExportFile.COLOUR_REQUIRED_SHEET].rows
    )
    for i, category_row in enumerate(category_rows):
        assert [row.value for row in worksheet_rows[i]] == category_row


@pytest.mark.django_db
def test_categories_sheet(categories, export_file):
    categories.sort(key=lambda x: x.english)
    category_rows = [[_.name, _.english, _.french] for _ in categories]
    worksheet_rows = list(
        export_file[_MissingInformationExportFile.CATEGORIES_SHEET].rows
    )
    assert [_.value for _ in worksheet_rows[0]] == ["Name", "English", "French"]
    for i, category_row in enumerate(category_rows, 1):
        assert [row.value for row in worksheet_rows[i]] == category_row


@pytest.mark.django_db
@pytest.mark.parametrize(
    "product_kwargs,expected_row",
    [
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 5.99,
                "english_size": "Large",
                "french_size__name": "Le Large",
                "colour": "Red",
                "fnac_range__category__name": "category name",
            },
            None,
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": None,
                "english_size": "Large",
                "french_size__name": "Le Large",
                "colour": "Red",
                "fnac_range__category__name": "category name",
            },
            [
                "389-738-478",
                "test product",
                None,
                "Large",
                "Le Large",
                "Red",
                f'=IF(COUNTIF({_MissingInformationExportFile.COLOUR_REQUIRED_SHEET}!A:C,H2),"Required","")',
                "category name",
            ],
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 599,
                "english_size": "Large",
                "french_size": None,
                "colour": "Red",
                "fnac_range__category__name": "category name",
            },
            [
                "389-738-478",
                "test product",
                "5.99",
                "Large",
                None,
                "Red",
                f'=IF(COUNTIF({_MissingInformationExportFile.COLOUR_REQUIRED_SHEET}!A:C,H2),"Required","")',
                "category name",
            ],
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 599,
                "english_size": "",
                "french_size": None,
                "colour": "Red",
                "fnac_range__category__name": "category name",
            },
            None,
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 599,
                "english_size": "",
                "french_size": None,
                "colour": "Red",
                "fnac_range__category": None,
            },
            [
                "389-738-478",
                "test product",
                "5.99",
                None,
                None,
                "Red",
                f'=IF(COUNTIF({_MissingInformationExportFile.COLOUR_REQUIRED_SHEET}!A:C,H2),"Required","")',
                None,
            ],
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": None,
                "english_size": "",
                "french_size": None,
                "colour": "Red",
                "fnac_range__category__name": "",
                "fnac_range__category__english": "category name",
            },
            [
                "389-738-478",
                "test product",
                None,
                None,
                None,
                "Red",
                f'=IF(COUNTIF({_MissingInformationExportFile.COLOUR_REQUIRED_SHEET}!A:C,H2),"Required","")',
                "category name",
            ],
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 599,
                "english_size": "",
                "french_size": None,
                "colour": "",
                "fnac_range__category__name": "",
                "fnac_range__category__english": "category name",
            },
            None,
        ),
        (
            {
                "sku": "389-738-478",
                "name": "test product",
                "price": 599,
                "english_size": "",
                "french_size": None,
                "colour": "",
                "fnac_range__category__name": "",
                "fnac_range__category__english": "category name",
                "fnac_range__category__requires_colour": True,
            },
            [
                "389-738-478",
                "test product",
                "5.99",
                None,
                None,
                None,
                f'=IF(COUNTIF({_MissingInformationExportFile.COLOUR_REQUIRED_SHEET}!A:C,H2),"Required","")',
                "category name",
            ],
        ),
    ],
)
def test_product_rows(product_kwargs, expected_row, fnac_product_factory):
    fnac_product_factory.create(**product_kwargs)
    workbook = openpyxl.load_workbook(create_add_missing_information_export())
    worksheet_rows = list(workbook[_MissingInformationExportFile.PRODUCTS_SHEET].rows)
    if expected_row is None:
        assert len(worksheet_rows) == 1
    else:
        assert [row.value for row in worksheet_rows[1]] == expected_row


def test_product_sheet_header(export_file):
    worksheet = export_file[_MissingInformationExportFile.PRODUCTS_SHEET]
    header = [_.value for _ in list(worksheet.rows)[0]]
    assert header == _MissingInformationExportFile.HEADER
