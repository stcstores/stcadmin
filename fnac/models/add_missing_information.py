"""Create an export file for adding missing product information."""

import io
from tempfile import NamedTemporaryFile

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import models, transaction
from django.db.models import Q
from django.utils import timezone

from .category import Category
from .fnac_product import FnacProduct
from .size import Size


class MissingInformationExportManager(models.Manager):
    """Manager for the MissingInformationExport model."""

    def is_in_progress(self):
        """Return True if there is an export being created, otherwise False."""
        return (
            self.get_queryset()
            .filter(status=MissingInformationExport.IN_PROGRESS)
            .exists()
        )

    def get_filename(self):
        """Return a filename for an export."""
        date_string = timezone.now().strftime("%Y-%m-%d")
        return f"fnac_missing_information_{date_string}.xlsx"

    def create_export(self):
        """Create a missing information export."""
        with transaction.atomic():
            if self.is_in_progress():
                raise MissingInformationExport.AlreadyInProgress()
            export = self.create()
        try:
            export_file = create_add_missing_information_export()
            export.export = SimpleUploadedFile(self.get_filename(), export_file.read())
        except Exception as e:
            export.status = export.ERROR
            export.save()
            raise e
        else:
            export.status = export.COMPLETE
        export.save()


class MissingInformationExport(models.Model):
    """Model for missing infomation export files."""

    class AlreadyInProgress(Exception):
        """Exception raised when an export is created with one already in progress."""

        def __init__(self, *args, **kwargs):
            """Raise the exception."""
            return super().__init__(self, "An export is already being created.")

    COMPLETE = "complete"
    ERROR = "error"
    IN_PROGRESS = "in_progress"
    STATUSES = ((COMPLETE, "Complete"), (ERROR, "Error"), (IN_PROGRESS, "In Progress"))

    timestamp = models.DateTimeField(auto_now_add=True)
    status = models.CharField(max_length=15, choices=STATUSES, default=IN_PROGRESS)
    export = models.FileField(upload_to="fnac/missing_informatio_exports/", null=True)

    objects = MissingInformationExportManager()


class MissingInformationImportManager(models.Manager):
    """Manager for the MissingInformationImport model."""

    def is_in_progress(self):
        """Return True if there is an export being created, otherwise False."""
        return (
            self.get_queryset()
            .filter(status=MissingInformationImport.IN_PROGRESS)
            .exists()
        )

    def get_filename(self):
        """Return a filename for an imported file."""
        date_string = timezone.now().strftime("%Y-%m-%d")
        return f"fnac_missing_information_import_{date_string}.xlsx"

    def create_import(self, import_file):
        """Create a missing information import."""
        with transaction.atomic():
            if self.is_in_progress():
                raise MissingInformationImport.AlreadyInProgress()
            import_object = self.create(
                import_file=SimpleUploadedFile(self.get_filename(), import_file.read())
            )
        try:
            import_missing_information(import_object.import_file.path)
        except Exception as e:
            import_object.status = import_object.ERROR
            import_object.save()
            raise e
        else:
            import_object.status = import_object.COMPLETE
        import_object.save()


class MissingInformationImport(models.Model):
    """Model for missing infomation import files."""

    class AlreadyInProgress(Exception):
        """Exception raised when an import is created with one already in progress."""

        def __init__(self, *args, **kwargs):
            """Raise the exception."""
            return super().__init__(self, "An import is already in progress.")

    COMPLETE = "complete"
    ERROR = "error"
    IN_PROGRESS = "in_progress"
    STATUSES = ((COMPLETE, "Complete"), (ERROR, "Error"), (IN_PROGRESS, "In Progress"))

    timestamp = models.DateTimeField(auto_now_add=True)
    status = models.CharField(max_length=15, choices=STATUSES, default=IN_PROGRESS)
    import_file = models.FileField(
        upload_to="fnac/missing_information_imports/", null=True
    )

    objects = MissingInformationImportManager()


def create_add_missing_information_export():
    """Create an export file for adding missin product information."""
    return _MissingInformationExportFile().create()


def import_missing_information(import_file):
    """Import product information from a missing information import file."""
    workbook = openpyxl.load_workbook(import_file)
    _ImportMissingInformationFile().import_product_data(workbook)


class _MissingInformationFile:
    SKU = "SKU"
    NAME = "Name"
    CATEGORY = "Category"
    SIZE_ENGLISH = "Size UK"
    SIZE_FRENCH = "Size FR"
    COLOUR = "Colour"
    COLOUR_REQUIRED = "Colour Required"
    PRICE = "Price"

    PRODUCTS_SHEET = "Products"
    CATEGORIES_SHEET = "Categories"
    SIZES_SHEET = "Sizes"
    INSTRUCTIONS_SHEET = "instructions"
    COLOUR_REQUIRED_SHEET = "colour_required"

    HEADER = [
        SKU,
        NAME,
        PRICE,
        SIZE_ENGLISH,
        SIZE_FRENCH,
        COLOUR,
        COLOUR_REQUIRED,
        CATEGORY,
    ]


class _MissingInformationExportFile(_MissingInformationFile):

    INSTRUCTIONS = {
        _MissingInformationFile.SKU: "The product's SKU. DO NOT ALTER.",
        _MissingInformationFile.NAME: (
            "The name of the product in Cloud Commerce, for reference only."
        ),
        _MissingInformationFile.PRICE: (
            "The price at which the product will be sold on FNAC."
        ),
        _MissingInformationFile.SIZE_ENGLISH: (
            "The size of the product according to Cloud Commerce. "
            "For reference only."
        ),
        _MissingInformationFile.SIZE_FRENCH: (
            "The size as shown on FNAC. Must be one of the values "
            f"in the {_MissingInformationFile.SIZES_SHEET} worksheet."
        ),
        _MissingInformationFile.COLOUR: (
            "The colour of the product. This is required for certain categories. "
            'If the colour is required the word "Required" will appear in the '
            f"{_MissingInformationFile.COLOUR_REQUIRED} field. Otherwise it can "
            "be left empty. Enter in English, it will be translated with the names "
            "and descriptions"
        ),
        _MissingInformationFile.COLOUR_REQUIRED: (
            f"If the category set in the {_MissingInformationFile.CATEGORY} column "
            'requires a colour to be set this field will contain the word "Required", '
            "otherwise it will be empty. Do not change the contents of this field."
        ),
        _MissingInformationFile.CATEGORY: (
            "Category information can be found in the "
            f"{_MissingInformationFile.CATEGORIES_SHEET} worksheet. "
            "Enter the any value from the appropriate row."
        ),
    }

    def create(self):
        workbook = openpyxl.Workbook()
        self.add_products(workbook)
        self.add_categories(workbook)
        self.add_sizes(workbook)
        self.add_instructions(workbook)
        self.add_colour_categories_worksheet(workbook)
        for worksheet in workbook.worksheets:
            self.set_column_width(worksheet)
        return self.workbook_to_bytes(workbook)

    def get_row_for_product(self, product, row_number):
        return [
            product.sku,
            product.name,
            self.get_price(product),
            product.english_size,
            self.get_french_size(product),
            product.colour,
            f'=IF(COUNTIF({self.COLOUR_REQUIRED_SHEET}!A:C,H{row_number}),"Required","")',
            self.get_category(product.fnac_range),
        ]

    def get_category(self, fnac_range):
        if fnac_range.category is None:
            return None
        if fnac_range.category.name:
            return fnac_range.category.name
        return fnac_range.category.english

    def get_price(self, product):
        if product.price is None:
            return None
        price_float = float(product.price) / 100
        return f"{price_float:.2f}"

    def get_french_size(self, product):
        if product.french_size is None:
            return None
        return str(product.french_size)

    def get_products(self):
        return FnacProduct.objects.missing_information()

    def add_products(self, workbook):
        worksheet = workbook.active
        worksheet.title = self.PRODUCTS_SHEET
        self.add_row(worksheet, 1, self.HEADER)
        rows = self.get_product_rows()
        for row_number, row in enumerate(rows, 2):
            self.add_row(worksheet, row_number, row)

    def get_product_rows(self):
        rows = []
        for row_number, product in enumerate(self.get_products(), 2):
            rows.append(self.get_row_for_product(product, row_number))
        return rows

    def get_categories(self):
        categories = Category.objects.all()
        category_rows = [
            [category.name or None, category.english, category.french]
            for category in categories
        ]
        return category_rows

    def add_categories(self, workbook):
        header = ["Name", "English", "French"]
        category_rows = self.get_categories()
        worksheet = workbook.create_sheet(self.CATEGORIES_SHEET)
        self.add_row(worksheet, 1, header)
        for row_number, row in enumerate(category_rows, 2):
            self.add_row(worksheet, row_number, row)

    def add_colour_categories_worksheet(self, workbook):
        worksheet = workbook.create_sheet(self.COLOUR_REQUIRED_SHEET)
        rows = [
            [_.name, _.english, _.french]
            for _ in Category.objects.filter(requires_colour=True)
        ]
        for row_number, row in enumerate(rows, 1):
            self.add_row(worksheet, row_number, row)

    def add_sizes(self, workbook):
        worksheet = workbook.create_sheet(self.SIZES_SHEET)
        sizes = Size.objects.all().order_by("name").values_list("name", flat=True)
        for row_number, size in enumerate(sizes, 1):
            self.add_row(worksheet, row_number, [size])

    def add_instructions(self, workbook):
        worksheet = workbook.create_sheet(self.INSTRUCTIONS_SHEET)
        header = ["Column", "Use"]
        rows = [[col, self.INSTRUCTIONS[col]] for col in self.HEADER]
        self.add_row(worksheet, 1, header)
        for row_number, row in enumerate(rows, 2):
            self.add_row(worksheet, row_number, row)

    def add_row(self, worksheet, row_number, row_data):
        for column_number, value in enumerate(row_data, 1):
            worksheet.cell(column=column_number, row=row_number, value=value)

    def workbook_to_bytes(self, workbook):
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return io.BytesIO(stream)

    def set_column_width(self, worksheet):
        for column_cells in worksheet.columns:
            length = max(len(cell.value or "") for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length


class _ImportMissingInformationFile(_MissingInformationFile):
    def import_product_data(self, workbook):
        worksheet = workbook[self.PRODUCTS_SHEET]
        for row in list(worksheet.rows)[1:]:
            row = [_.value for _ in row]
            kwargs = self.get_update_kwargs(row)
            category = kwargs.pop("category", None)
            qs = FnacProduct.objects.filter(sku=row[self.column_number(self.SKU)])
            qs.update(**kwargs)
            if category is not None:
                try:
                    fnac_range = qs[0].fnac_range
                except IndexError:
                    return
                else:
                    fnac_range.category = category
                    fnac_range.save()

    def column_number(self, column_name):
        return self.HEADER.index(column_name)

    def get_update_kwargs(self, row):
        kwargs = {
            "french_size": self.get_update_size(row),
            "category": self.get_update_cateogry(row),
            "price": self.get_update_price(row),
            "colour": self.get_update_colour(row),
        }
        return {key: value for key, value in kwargs.items() if value}

    def get_update_size(self, row):
        text = row[self.column_number(self.SIZE_FRENCH)]
        try:
            return Size.objects.get(name=text)
        except Size.DoesNotExist:
            return None

    def get_update_cateogry(self, row):
        text = row[self.column_number(self.CATEGORY)]
        try:
            return Category.objects.get(Q(name=text) | Q(english=text) | Q(french=text))
        except Category.DoesNotExist:
            return None

    def get_update_price(self, row):
        text = row[self.column_number(self.PRICE)]
        if text:
            return int(float(text) * 100)

    def get_update_colour(self, row):
        return row[self.column_number(self.COLOUR)]
