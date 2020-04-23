"""Create an import file to create new products in FNAC."""

import io
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import models, transaction
from django.utils import timezone

from .fnac_product import FnacProduct


class NewProductExportManager(models.Manager):
    """Manager for the NewProductExport model."""

    def is_in_progress(self):
        """Return True if there is an export being created, otherwise False."""
        return self.get_queryset().filter(status=NewProductExport.IN_PROGRESS).exists()

    def get_filename(self):
        """Return a filename for an export."""
        date_string = timezone.now().strftime("%Y-%m-%d")
        return f"fnac_new_products_{date_string}.xlsx"

    def create_export(self):
        """Create a missing information export."""
        with transaction.atomic():
            if self.is_in_progress():
                raise NewProductExport.AlreadyInProgress()
            export = self.create()
        try:
            export_file = create_new_product_export()
            export.export = SimpleUploadedFile(self.get_filename(), export_file.read())
        except Exception as e:
            export.status = export.ERROR
            export.save()
            raise e
        else:
            export.status = export.COMPLETE
        export.save()


class NewProductExport(models.Model):
    """Model for new product export files."""

    class AlreadyInProgress(Exception):
        """Exception raised when an export is created with one already in progress."""

        def __init__(self, *args, **kwargs):
            """Raise the exception."""
            return super().__init__(self, "An export is already being created.")

    COMPLETE = "complete"
    ERROR = "error"
    IN_PROGRESS = "in_progress"
    STATUSES = ((COMPLETE, "Complete"), (ERROR, "Error"), (IN_PROGRESS, "In Progress"))

    timestamp = models.DateTimeField(auto_now_add=True)
    status = models.CharField(max_length=15, choices=STATUSES, default=IN_PROGRESS)
    export = models.FileField(upload_to="fnac/new_product_files/", null=True)

    objects = NewProductExportManager()


def create_new_product_export():
    """Return a new product upload file."""
    return _ProductUpload().create()


class _ProductUpload:
    TEMPLATE_PATH = Path(__file__).parent / "new_product_export_template.xlsx"

    NAME = "Libellé"
    DESCRIPTION = "Description"
    CATEGORY = "Typologie"
    SKU = "SKU"
    BARCODE = "EAN"
    BRAND = "Constructeur Vendeur"
    COLOUR = "Coloris"
    SIZE = "Taille"
    IMAGES = [
        "Visuel Principal",
        "Visuel Secondaire 1",
        "Visuel Secondaire 2",
        "Visuel Secondaire 3",
    ]
    FORMAT = "Format"
    ADDITIONAL_DESCRIPTION = "Description complémentaire"
    COOKING_ACCESSORIES = "GRP_Cooking_accessories_type/attributeValue"
    TABLEWARE_TYPE = "GRP_Tableware_type/attributeValue"

    formats = {
        "Accessoires": ["accessory"],
        "Accessoires de bain": ["bath"],
        "Autocollant / Magnets": ["sticker", "magnet"],
        "Badge": ["badge"],
        "Bijoux": ["jewel"],
        "Blouson": ["jacket"],
        "Boucle de ceinture": ["buckle"],
        "Casquette": ["cap"],
        "Ceinture": ["belt"],
        "Chapeau": ["hat"],
        "Chaussures": ["shoe"],
        "Chemise": ["shirt"],
        "Cravate": ["tie"],
        "Drap de plage": ["towel"],
        "Echarpe": ["scarf"],
        "Ecusson": ["patch"],
        "Grande Figurine": ["figurine"],
        "Jeu de cartes": ["card"],
        "Objet de décoration": ["decorative"],
        "Pantalon": ["trouser"],
        "Parures de lit": ["bed"],
        "Peluche": ["plush"],
        "Pendentif": ["pendant", "necklace"],
        "Pendule et horloge": ["clock"],
        "Pins": ["pin"],
        "Porte-clef": ["keychain", "keyring"],
        "Portefeuille": ["wallet"],
        "Robe": ["dress"],
        "Sac": ["bag"],
        "Short": ["shorts"],
        "Sweatshirt": ["sweatshirt", "jumper", "cardigan"],
        "T-Shirt": ["tshirt", "t-shirt", "t shirt"],
        "Vaisselle": ["washing"],
        "Vêtement bébé": ["baby"],
        "Vêtement enfant": ["child", "kid"],
        "Voiture / Bateau": ["car", "boat"],
    }

    def get_row_for_product(self, product):
        if product.french_size is not None:
            size = product.french_size.name
        else:
            size = ""
        return {
            self.SKU: product.sku,
            self.NAME: product.translation.name,
            self.DESCRIPTION: product.translation.description,
            self.ADDITIONAL_DESCRIPTION: product.translation.description,
            self.FORMAT: self.get_format(product),
            self.CATEGORY: product.fnac_range.category.french,
            self.BARCODE: product.barcode,
            self.BRAND: product.brand,
            self.COLOUR: product.translation.colour,
            self.SIZE: size,
            self.IMAGES[0]: self.image_uri(product.image_1),
            self.IMAGES[1]: self.image_uri(product.image_2),
            self.IMAGES[2]: self.image_uri(product.image_3),
            self.IMAGES[3]: self.image_uri(product.image_4),
        }

    def image_uri(self, image_path):
        if image_path:
            return self.image_url + image_path
        return ""

    def add_row(self, worksheet, row_number, row_data):
        for key, value in row_data.items():
            column_number = self.header[0].index(key) + 1
            worksheet.cell(column=column_number, row=row_number, value=value)

    def get_products(self):
        return FnacProduct.objects.valid_information()

    def get_format(self, product):
        for format_string, match_words in self.formats.items():
            for word in match_words:
                if word in product.name:
                    return format_string
        return "Accessoires"

    def worksheet_row(self, worksheet, row_number):
        return [_.value for _ in list(worksheet.rows)[row_number]]

    def add_header(self, worksheet, header):
        for row_number, row in enumerate(header, 1):
            for column_number, value in enumerate(row, 1):
                worksheet.cell(column=column_number, row=row_number, value=value)

    def create(self):
        template = openpyxl.load_workbook(self.TEMPLATE_PATH)
        self.image_url = settings.CC_IMAGE_URL
        self.header = [
            self.worksheet_row(template.active, 0),
            self.worksheet_row(template.active, 1),
        ]
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        self.add_header(worksheet, self.header)
        products = self.get_products()
        for row_number, product in enumerate(products, len(self.header) + 1):
            row_data = self.get_row_for_product(product)
            self.add_row(worksheet, row_number, row_data)
        return self.workbook_to_bytes(workbook)

    def workbook_to_bytes(self, workbook):
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return io.BytesIO(stream)
