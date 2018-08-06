"""FileSecuredMailManifest class."""

import datetime
import io
import os

import openpyxl
from django.contrib import messages
from django.core.files import File
from spring_manifest import models

from .file_manifest import FileManifest


class FileSecuredMailManifest(FileManifest):
    """File a Secured Mail manifest."""

    PROOF_OF_DELIVERY = {"SMIU": "", "SMIT": "S"}
    SERVICE = {"SMIU": "International Untracked", "SMIT": "International Tracked"}
    TRACKED = "Tracked"

    def process_manifest(self):
        """File manifest."""
        self.orders = self.manifest.manifestorder_set.all()
        if self.valid():
            self.save_manifest_file()
        if self.valid():
            self.save_docket_file()
        if self.valid():
            self.manifest.status = self.manifest.FILED
            self.manifest.save()
            self.cleanup()
        else:
            self.manifest.time_filed = None
            self.manifest.manifest_file = None
            self.manifest.status = self.manifest.FAILED
            self.manifest.save()

    def invalid_country_message(self, order):
        """Return message for invalid countries."""
        return "Order {}: Country {} info invalid.".format(order, order.country)

    @staticmethod
    def get_packages_for_orders(orders):
        """Return a list of packages belonging to the passed orders."""
        return sum([list(order.packages) for order in orders], [])

    def save_manifest_file(self):
        """Create Manifest file and save to database."""
        manifest_file = SecuredMailManifestFile(self.manifest, self.orders)
        self.manifest.manifest_file.save(
            str(self.manifest) + "_manifest.xlsx", File(manifest_file)
        )

    def save_docket_file(self):
        """Create Docket file and save to database."""
        docket_orders = [
            order for order in self.orders if order.secured_mail_service.on_docket
        ]
        packages = self.get_packages_for_orders(docket_orders)
        docket_file = SecuredMailDocketFile(self.manifest, packages)
        self.manifest.docket_file.save(
            str(self.manifest) + "_docket.xlsx", File(docket_file)
        )
        self.manifest.file_manifest()

    def get_date_string(self):
        """Return currenct date as string."""
        return datetime.datetime.now().strftime("%Y-%m-%d")

    def add_success_messages(self, manifest):
        """Create success message."""
        orders = manifest.manifestorder_set.all()
        package_count = sum(o.manifestpackage_set.count() for o in orders)
        order_count = len(orders)
        messages.add_message(
            self.request,
            messages.SUCCESS,
            "Manifest file created with {} packages for {} orders".format(
                package_count, order_count
            ),
        )

    def cleanup(self):
        """Increase docket number."""
        self.increment_docket_number()

    def increment_docket_number(self):
        """Update Docket number in database."""
        counter = models.Counter.objects.get(name="Secured Mail Docket Number")
        counter.count += 1
        counter.save()


class SecuredMailManifestFile:
    """Create a Secured Mail Manifest file."""

    TEMPLATE_FILENAME = "secure_mail_manifest_template.xlsx"
    TEMPLATE_PATH = os.path.join(
        os.path.abspath(os.path.dirname(__file__)), TEMPLATE_FILENAME
    )
    ITEM_COL = "M"
    WEIGHT_COL = "N"
    REFERENCE_CELL = "J3"
    DATE_CELL = "O3"
    TRACKED_ROW = "22"

    def __new__(self, manifest, orders):
        """Create Secured Mail Manifest file."""
        untracked_service = models.ManifestService.objects.get(
            name="Secured Mail International Untracked"
        )
        tracked_service = models.ManifestService.objects.get(
            name="Secured Mail International Tracked"
        )
        untracked_orders = [_ for _ in orders if _.service == untracked_service]
        tracked_orders = [_ for _ in orders if _.service == tracked_service]
        tracked_weight = self.convert_weight(
            self, sum([o.weight for o in tracked_orders])
        )
        destinations = models.SecuredMailDestination.objects.all()
        wb = openpyxl.load_workbook(filename=self.TEMPLATE_PATH)
        ws = wb.active
        total_items = 0
        total_weight = 0
        for destination in destinations:
            orders_for_destination = [
                order
                for order in untracked_orders
                if order.country.secured_mail_destination == destination
            ]
            weight_for_destination = self.convert_weight(
                self, sum([o.weight for o in orders_for_destination])
            )
            total_items += len(orders_for_destination)
            total_weight += weight_for_destination
            row = str(destination.manifest_row_number)
            ws[self.ITEM_COL + row] = len(orders_for_destination)
            ws[self.WEIGHT_COL + row] = weight_for_destination
        ws[self.ITEM_COL + self.TRACKED_ROW] = len(tracked_orders)
        ws[self.WEIGHT_COL + self.TRACKED_ROW] = tracked_weight
        ws[self.REFERENCE_CELL] = str(manifest)
        ws[self.DATE_CELL] = datetime.datetime.now().strftime("%d/%m/%Y")
        return io.BytesIO(openpyxl.writer.excel.save_virtual_workbook(wb))

    def convert_weight(self, weight):
        """Return a gram weight as kilograms."""
        return round(weight / 1000, 2)


class SecuredMailDocketFile:
    """Create Secured Docket file."""

    TEMPLATE_FILENAME = "secure_mail_docket_template.xlsx"
    TEMPLATE_PATH = os.path.join(
        os.path.abspath(os.path.dirname(__file__)), TEMPLATE_FILENAME
    )
    COLLECTION_SITE_NAME = "Seaton Trading Company Ltd"
    CONTACT_NAME = "Larry Guogis"
    CONTACT_NUMBER = "01297 21874"
    CONTACT_ADDRESS = "26 Harbour Road, Seaton, EX12 2NA"
    CLIENT_TO_BE_BILLED = "Seaton Trading Company Ltd"
    PRINTED_NAME = "Larry Guogis"
    INITIALS = "ST"
    JOB_NAME = INITIALS
    COLLECTION_SITE_CELL = "D3"
    DOCKET_NUMBER_CELL = "F3"
    COLLECTION_DATE_CELL = "I3"
    CONTACT_NAME_CELL = "D6"
    CONTACT_NUMBER_CELL = "I6"
    CONTACT_ADDRESS_CELL = "D8"
    CLIENT_TO_BE_BILLED_CELL = "D10"
    PRINTED_NAME_CELL = "C34"
    SIGN_DATE_FIELD = "C36"
    TABLE_START_ROW = 16
    JOB_NAME_COL = "B"
    SERVICE_COL = "C"
    FORMAT_COL = "E"
    ITEM_WEIGHT_COL = "F"
    QUANTITY_MAILED_COL = "G"
    PRESENTATION_TYPE_COL = "H"
    PRESENTATION_QUANTITY_COL = "I"

    def __new__(self, manifest, packages):
        """Create Secured Docket file."""
        self.workbook = openpyxl.load_workbook(filename=self.TEMPLATE_PATH)
        self.worksheet = self.workbook.active
        self.packages = packages
        self.write_worksheet(self)
        return io.BytesIO(openpyxl.writer.excel.save_virtual_workbook(self.workbook))

    def write_worksheet(self):
        """Write data to the worksheet."""
        self.fill_form(self)
        services = models.SecuredMailService.objects.filter(on_docket=True)
        for row_number, service in enumerate(services):
            self.write_service_row(self, row_number, service)

    def fill_form(self):
        """Fill out form fields in the worksheet."""
        ws = self.worksheet
        ws[self.COLLECTION_SITE_CELL] = self.COLLECTION_SITE_NAME
        ws[self.DOCKET_NUMBER_CELL] = self.get_docket_number(self)
        ws[self.CONTACT_NAME_CELL] = self.CONTACT_NAME
        ws[self.CONTACT_ADDRESS_CELL] = self.CONTACT_ADDRESS
        ws[self.COLLECTION_DATE_CELL] = self.get_date()
        ws[self.CONTACT_NUMBER_CELL] = self.CONTACT_NUMBER
        ws[self.CLIENT_TO_BE_BILLED_CELL] = self.CLIENT_TO_BE_BILLED
        ws[self.PRINTED_NAME_CELL] = self.PRINTED_NAME
        ws[self.SIGN_DATE_FIELD] = self.get_date()

    def write_service_row(self, row_number, service):
        """Add a service row to the worksheet."""
        ws = self.worksheet
        sheet_row = str(self.TABLE_START_ROW + row_number)
        packages = [_ for _ in self.packages if _.order.secured_mail_service == service]
        if len(packages) == 0:
            return
        weights = [
            _.weight for _ in self.packages if _.order.secured_mail_service == service
        ]
        ws[self.JOB_NAME_COL + sheet_row] = self.JOB_NAME
        ws[self.SERVICE_COL + sheet_row] = service.docket_service
        ws[self.FORMAT_COL + sheet_row] = service.format
        ws[self.ITEM_WEIGHT_COL + sheet_row] = int(sum(weights) / len(weights))
        ws[self.QUANTITY_MAILED_COL + sheet_row] = len(weights)

    def get_docket_number(self):
        """Return current docket number."""
        counter = models.Counter.objects.get(name="Secured Mail Docket Number")
        docket_number = counter.count
        return f"{self.INITIALS}{self.INITIALS}{docket_number}"

    @staticmethod
    def get_date():
        """Return current date as string."""
        return datetime.datetime.now().strftime("%d/%m/%Y")
